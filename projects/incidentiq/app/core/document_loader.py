"""Load incident knowledge from PDF, DOCX, TXT, and XLSX files on disk.

Used at build time by ``scripts/build_knowledge_base.py`` to ingest files from
``data/documents/``. Each file is normalised into a dict with ``id``, ``title``,
``source_file``, ``document_type``, ``category``, ``severity``, ``tags``, and
plain ``text`` suitable for chunking and embedding.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from app.utils.logger import get_logger

_logger = get_logger(__name__)

_SUPPORTED_EXTENSIONS: frozenset[str] = frozenset({".txt", ".pdf", ".docx", ".xlsx"})
_ID_PATTERN: re.Pattern[str] = re.compile(
    r"^(INC-\d+|SOP-[A-Z0-9-]+|REF-\d+)",
    re.IGNORECASE,
)
_EXCEL_HEADERS: frozenset[str] = frozenset(
    {"id", "title", "severity", "category", "description"}
)


def _project_root() -> Path:
    return Path(__file__).resolve().parent.parent.parent


def default_documents_dir() -> Path:
    """Return the canonical raw-documents directory."""
    return _project_root() / "data" / "documents"


def _derive_id(stem: str) -> str:
    """Derive a stable document id from a filename stem."""
    base = stem.split("_")[0]
    normalised = base.replace(" ", "_")
    match = _ID_PATTERN.match(normalised.replace("_", "-"))
    if match:
        return match.group(1).upper()
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", stem).strip("-").lower()
    return slug or "document"


def _derive_title(stem: str, text: str) -> str:
    """Best-effort title from filename or first line of extracted text."""
    cleaned = stem.replace("_", " ").replace("-", " ")
    if cleaned and not cleaned.lower().startswith("incident register"):
        return cleaned.title()
    first_line = next((line.strip() for line in text.splitlines() if line.strip()), "")
    if len(first_line) <= 120:
        return first_line or stem
    return first_line[:117] + "..."


def _infer_document_type(stem: str, text: str) -> str:
    upper = stem.upper()
    if upper.startswith("SOP") or " SOP " in text[:200].upper():
        return "sop"
    if upper.startswith("REF") or "REFERENCE" in text[:200].upper():
        return "reference"
    return "incident"


def _infer_severity(stem: str, text: str) -> str:
    for token in ("P1", "P2", "P3"):
        if token in stem.upper() or token in text[:300].upper():
            return token
    return "N/A"


def _infer_category(stem: str, text: str) -> str:
    lower = (stem + " " + text[:400]).lower()
    mapping = {
        "postgres": "Database",
        "mysql": "Database",
        "mongo": "Database",
        "kafka": "MessageQueue",
        "rabbit": "MessageQueue",
        "kubernetes": "Kubernetes",
        "pod": "Kubernetes",
        "crashloop": "Kubernetes",
        "network": "Network",
        "dns": "Network",
        "vpn": "Network",
        "api gateway": "Application",
        "redis": "Application",
        "aws": "Cloud",
        "azure": "Cloud",
        "gcp": "Cloud",
    }
    for needle, category in mapping.items():
        if needle in lower:
            return category
    return "Reference" if stem.lower().startswith("incident_register") else "Unknown"


def _read_txt(path: Path) -> str:
    return path.read_text(encoding="utf-8").strip()


def _read_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n".join(pages).strip()


def _read_docx(path: Path) -> str:
    from docx import Document

    doc = Document(str(path))
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs).strip()


def _read_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(str(path), read_only=True, data_only=True)
    blocks: list[str] = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
        has_schema = _EXCEL_HEADERS.issubset(set(header))
        blocks.append(f"Sheet: {sheet.title}")
        if has_schema:
            col = {name: header.index(name) for name in _EXCEL_HEADERS}
            for row in rows[1:]:
                if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                    continue
                values = [row[i] if i < len(row) else "" for i in col.values()]
                row_id, title, severity, category, description = values
                blocks.append(
                    f"Incident {row_id} [{severity} - {category}]: {title}. "
                    f"Description: {description}."
                )
        else:
            for row in rows:
                cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if cells:
                    blocks.append(" | ".join(cells))
    workbook.close()
    return "\n".join(blocks).strip()


def extract_text(path: Path) -> str:
    """Extract plain text from a supported document path."""
    suffix = path.suffix.lower()
    if suffix == ".txt":
        return _read_txt(path)
    if suffix == ".pdf":
        return _read_pdf(path)
    if suffix == ".docx":
        return _read_docx(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    raise ValueError(f"Unsupported extension: {suffix}")


def load_documents_from_folder(folder: Path | str | None = None) -> list[dict[str, Any]]:
    """Scan ``folder`` recursively and return normalised document records.

    Args:
        folder: Directory to scan. Defaults to ``data/documents/``.

    Returns:
        List of dicts with keys ``id``, ``title``, ``source_file``, ``document_type``,
        ``category``, ``severity``, ``tags``, and ``text``. Returns ``[]`` when the
        folder is missing or empty.
    """
    root = Path(folder) if folder is not None else default_documents_dir()
    if not root.is_dir():
        _logger.warning("Documents folder not found, skipping file ingestion: path=%s", root)
        return []

    records: list[dict[str, Any]] = []
    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        suffix = path.suffix.lower()
        if suffix not in _SUPPORTED_EXTENSIONS:
            _logger.warning("Skipping unsupported file type: path=%s", path.name)
            continue
        try:
            text = extract_text(path)
        except Exception as exc:
            _logger.error("Failed to read document: path=%s error=%s", path.name, exc)
            continue
        if not text:
            _logger.warning("Skipping empty document: path=%s", path.name)
            continue

        stem = path.stem
        doc_id = _derive_id(stem)
        doc_type = _infer_document_type(stem, text)
        record: dict[str, Any] = {
            "id": doc_id,
            "title": _derive_title(stem, text),
            "source_file": path.name,
            "document_type": doc_type,
            "category": _infer_category(stem, text),
            "severity": _infer_severity(stem, text),
            "tags": [suffix.lstrip(".")],
            "text": text,
        }
        records.append(record)
        _logger.info(
            "Loaded file document: id=%s source=%s chars=%d",
            doc_id,
            path.name,
            len(text),
        )

    _logger.info("File ingestion complete: folder=%s documents=%d", root, len(records))
    return records
